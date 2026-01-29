"""
Excel Handler Module.
Reads and writes Excel files while preserving data and adding new columns.
"""

from pathlib import Path
from typing import List, Dict, Optional, Any
from datetime import datetime
import openpyxl
from openpyxl import Workbook
import shutil

from ..models import Lead, EmailVerificationStatus, SequenceStep
from ..config import DATA_DIR


# Expected columns from the original Excel
ORIGINAL_COLUMNS = [
    'name', 'email', 'website', 'category', 'city'
]

# New columns we add
NEW_COLUMNS = [
    'email_verified', 'verification_checked_at', 'website_scan_summary',
    'website_scan_at', 'my_notes', 'email_subject', 'email_draft', 
    'email_sent_at', 'owner_name', 'sequence_step', 'their_last_reply', 
    'my_reply_draft'
]


class ExcelHandler:
    """Handles reading and writing Excel files with lead data."""
    
    def __init__(self, file_path: Path):
        self.file_path = Path(file_path)
        self.workbook: Optional[Workbook] = None
        self.leads: List[Lead] = []
        self.column_mapping: Dict[str, int] = {}
        self.all_columns: List[str] = []
    
    def load(self) -> List[Lead]:
        """Load leads from Excel or CSV file."""
        if self.file_path.suffix.lower() == '.csv':
            return self._load_csv()
        else:
            return self._load_excel()

    def _load_excel(self) -> List[Lead]:
        """Load leads from Excel file."""
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True, read_only=True)
        sheet = self.workbook.active
        
        # Get all rows generator
        rows = sheet.iter_rows(values_only=True)
        
        # Get headers from first row
        try:
            header_row = next(rows)
        except StopIteration:
            return [] # Empty file
            
        headers = []
        for idx, cell_value in enumerate(header_row):
            if cell_value:
                headers.append(str(cell_value).lower().strip())
            else:
                headers.append(f"column_{idx+1}")
        
        self.all_columns = headers.copy()
        self.column_mapping = {header: idx for idx, header in enumerate(headers)}
        
        # Add new columns if they don't exist
        for new_col in NEW_COLUMNS:
            if new_col not in self.column_mapping:
                self.all_columns.append(new_col)
                self.column_mapping[new_col] = len(self.all_columns) - 1
        
        # Parse rows into Lead objects
        self.leads = []
        # enumerate starts at 2 because we skipped header (idx 1)
        for row_idx, row_values in enumerate(rows, 2):
            # Row values is a tuple of values
            
            # Skip empty rows (all None or empty strings)
            if not any(v is not None and str(v).strip() != "" for v in row_values):
                continue
            
            row_data = {}
            for i, val in enumerate(row_values):
                if i < len(headers):
                    row_data[headers[i]] = val
            
            lead = self._row_to_lead(row_idx, row_data)
            self.leads.append(lead)
        
        # Close the workbook as we are done reading
        self.workbook.close()
        
        return self.leads

    def _load_csv(self) -> List[Lead]:
        """Load leads from CSV file."""
        import csv
        
        self.leads = []
        try:
            with open(self.file_path, mode='r', encoding='utf-8-sig', errors='replace') as f:
                reader = csv.DictReader(f)
                headers = [h.lower().strip() for h in reader.fieldnames]
                # Re-map reader headers to lower case for consistency
                reader.fieldnames = headers
                
                self.all_columns = headers.copy()
                self.column_mapping = {header: idx for idx, header in enumerate(headers)}
                
                # Add new columns if they don't exist
                for new_col in NEW_COLUMNS:
                    if new_col not in self.column_mapping:
                        self.all_columns.append(new_col)
                        self.column_mapping[new_col] = len(self.all_columns) - 1
                
                for i, row in enumerate(reader, 1):
                    # Skip empty rows
                    if not any(row.values()):
                        continue
                    lead = self._row_to_lead(i, row)
                    self.leads.append(lead)
        except Exception as e:
            print(f"Error loading CSV: {e}")
            raise
            
        return self.leads

    def save(self, output_path: Path = None) -> Path:
        """Save leads back to Excel or CSV file."""
        if output_path is None:
            output_path = self.file_path
        
        if output_path.suffix.lower() == '.csv':
            return self._save_csv(output_path)
        else:
            return self._save_excel(output_path)

    def _save_excel(self, output_path: Path) -> Path:
        """Save leads back to Excel file."""
        # Create a new workbook
        wb = Workbook()
        sheet = wb.active
        
        # Write headers
        for col_idx, col_name in enumerate(self.all_columns, 1):
            sheet.cell(row=1, column=col_idx, value=col_name)
        
        # Write data
        for row_idx, lead in enumerate(self.leads, 2):
            lead_dict = self._lead_to_row(lead)
            for col_name, col_idx in self.column_mapping.items():
                value = lead_dict.get(col_name)
                sheet.cell(row=row_idx, column=col_idx + 1, value=value)
        
        wb.save(output_path)
        return output_path

    def _save_csv(self, output_path: Path) -> Path:
        """Save leads back to CSV file."""
        import csv
        with open(output_path, mode='w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=self.all_columns)
            writer.writeheader()
            for lead in self.leads:
                writer.writerow(self._lead_to_row(lead))
        return output_path

    def _row_to_lead(self, row_id: int, row_data: Dict[str, Any]) -> Lead:
        """Convert a row dict to a Lead object."""
        # Extract known fields
        name = str(row_data.get('name', '') or '')
        email = str(row_data.get('email', '') or '')
        website = row_data.get('website')
        category = row_data.get('category')
        city = row_data.get('city')
        
        # Parse verification status
        email_verified_str = str(row_data.get('email_verified', '') or '').lower()
        try:
            email_verified = EmailVerificationStatus(email_verified_str)
        except ValueError:
            email_verified = EmailVerificationStatus.PENDING
        
        # Parse sequence step
        sequence_step_str = str(row_data.get('sequence_step', '') or '').lower()
        try:
            sequence_step = SequenceStep(sequence_step_str)
        except ValueError:
            sequence_step = SequenceStep.NOT_SENT
        
        # Parse datetime fields
        verification_checked_at = self._parse_datetime(row_data.get('verification_checked_at'))
        website_scan_at = self._parse_datetime(row_data.get('website_scan_at'))
        email_sent_at = self._parse_datetime(row_data.get('email_sent_at'))
        
        # Collect extra data (columns we don't explicitly handle)
        extra_data = {}
        known_fields = set(ORIGINAL_COLUMNS + NEW_COLUMNS)
        for key, value in row_data.items():
            if key not in known_fields and value is not None:
                extra_data[key] = value
        
        return Lead(
            id=row_id,
            name=name,
            email=email,
            website=str(website) if website else None,
            category=str(category) if category else None,
            city=str(city) if city else None,
            email_verified=email_verified,
            verification_checked_at=verification_checked_at,
            website_scan_summary=row_data.get('website_scan_summary'),
            website_scan_at=website_scan_at,
            my_notes=row_data.get('my_notes'),
            email_subject=row_data.get('email_subject'),
            email_draft=row_data.get('email_draft'),
            email_sent_at=email_sent_at,
            owner_name=row_data.get('owner_name') or name,
            sequence_step=sequence_step,
            their_last_reply=row_data.get('their_last_reply'),
            my_reply_draft=row_data.get('my_reply_draft'),
            extra_data=extra_data
        )

    def _parse_datetime(self, value: Any) -> Optional[datetime]:
        """Parse datetime from various formats."""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value
        try:
            return datetime.fromisoformat(str(value))
        except:
            return None

    def _lead_to_row(self, lead: Lead) -> Dict[str, Any]:
        """Convert a Lead object back to a row dict."""
        row = {
            'name': lead.name,
            'email': lead.email,
            'website': lead.website,
            'category': lead.category,
            'city': lead.city,
            'email_verified': lead.email_verified.value,
            'verification_checked_at': lead.verification_checked_at.isoformat() if lead.verification_checked_at else None,
            'website_scan_summary': lead.website_scan_summary,
            'website_scan_at': lead.website_scan_at.isoformat() if lead.website_scan_at else None,
            'my_notes': lead.my_notes,
            'email_subject': lead.email_subject,
            'email_draft': lead.email_draft,
            'email_sent_at': lead.email_sent_at.isoformat() if lead.email_sent_at else None,
            'owner_name': lead.owner_name,
            'sequence_step': lead.sequence_step.value,
            'their_last_reply': lead.their_last_reply,
            'my_reply_draft': lead.my_reply_draft,
        }

        
        # Add extra data
        row.update(lead.extra_data)
        
        return row
    
    def update_lead(self, lead_id: int, updates: Dict[str, Any]) -> Optional[Lead]:
        """Update a specific lead by ID."""
        for i, lead in enumerate(self.leads):
            if lead.id == lead_id:
                lead_dict = lead.model_dump()
                lead_dict.update(updates)
                self.leads[i] = Lead(**lead_dict)
                return self.leads[i]
        return None
    
    def get_lead(self, lead_id: int) -> Optional[Lead]:
        """Get a specific lead by ID."""
        for lead in self.leads:
            if lead.id == lead_id:
                return lead
        return None
    
    def backup(self) -> Path:
        """Create a backup of the current file."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"{self.file_path.stem}_backup_{timestamp}{self.file_path.suffix}"
        backup_path = DATA_DIR / backup_name
        shutil.copy2(self.file_path, backup_path)
        return backup_path


# Global instance for the current working file
_current_handler: Optional[ExcelHandler] = None


def get_handler() -> Optional[ExcelHandler]:
    """Get the current Excel handler instance."""
    return _current_handler


def set_handler(handler: ExcelHandler) -> None:
    """Set the current Excel handler instance."""
    global _current_handler
    _current_handler = handler


def load_excel(file_path: Path) -> ExcelHandler:
    """Load an Excel file and set it as the current handler."""
    handler = ExcelHandler(file_path)
    handler.load()
    set_handler(handler)
    return handler
